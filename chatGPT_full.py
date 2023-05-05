import openpyxl
import time
import openai

openai.api_key = "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"


def generate_answer(question):
    response = openai.Completion.create(
        engine="davinci",
        prompt=question,
        max_tokens=1024,
        n=1,
        stop=None,
        temperature=0.5,
    )
    answer = response.choices[0].text.strip()
    return answer


def answer_questions(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row_num in range(1, ws.max_row + 1):
        question_cell = ws.cell(row=row_num, column=1)
        answer_cell = ws.cell(row=row_num, column=2)

        if not question_cell.value:
            # Skip empty rows
            continue

        question = question_cell.value
        print(question)

        # Generate the answer using ChatGPT
        answer = generate_answer(question)

        # Write the answer to the answer cell in the same row
        answer_cell.value = answer

        # Pause for a short time to avoid rate limiting
        time.sleep(0.5)

    # Save the updated Excel file
    wb.save(excel_file)


# reading the excel file
excel_file = "Question.xlsx"
# Answer the questions in the Excel file
answer_questions(excel_file)
